from pathlib import Path
from typing import List

from openpyxl import load_workbook

from .models import MasterItem, ArrivalItem


class Loader:
    """Loads MASTER and ARRIVAL Excel files."""

    def load_master(self, filename: str | Path) -> List[MasterItem]:
        wb = load_workbook(filename=filename, data_only=True)
        ws = wb.active

        items: List[MasterItem] = []

        # Expected columns:
        # A=SKU B=Category C=Brand D=Variant E=Volume
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue

            items.append(
                MasterItem(
                    sku=str(row[0]).strip(),
                    category=str(row[1] or "").strip(),
                    brand=str(row[2] or "").strip(),
                    variant=str(row[3] or "").strip(),
                    volume=str(row[4] or "").strip(),
                )
            )

        return items

    def load_arrival(self, filename: str | Path) -> List[ArrivalItem]:
        wb = load_workbook(filename=filename, data_only=True)
        ws = wb.active

        items: List[ArrivalItem] = []

        # Column A = НАИМЕНОВАНИЕ
        for excel_row, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            name = row[0]

            if name is None:
                continue

            text = str(name).strip()

            if not text:
                continue

            items.append(
                ArrivalItem(
                    row_number=excel_row,
                    source_name=text,
                )
            )

        return items
